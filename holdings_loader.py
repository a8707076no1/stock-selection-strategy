"""
從「資產與持股明細更新案夾」自動載入最新持股清單。

xlsx 格式（第 1 列為標頭）：
  股票代號 | 公司名稱 | 持股數量 | 成本價 (TWD) | 備註

支援：
- 持股數量：「20.3 張」「110股」「3張」「5.3」等多種寫法
- 成本價：可空白（None）
- ETF 自動辨識：00 開頭代碼
- 同一目錄會挑「檔名上日期最新」的檔案
"""
import os, re, glob
from datetime import datetime

BASE_DIR = (os.environ.get("STOCK_BASE_DIR") or os.path.expanduser("~/Desktop/Stock Selection Strategy"))
HOLDINGS_FOLDER = os.path.join(BASE_DIR, "資產與持股明細更新案夾")

# DEFAULT_HOLDINGS 為空 — 真實持股由 資產與持股明細更新案夾/*.xlsx 提供（gitignore 排除）
DEFAULT_HOLDINGS = []


def find_latest_holdings_xlsx(folder=HOLDINGS_FOLDER):
    """從資料夾找檔名上日期最新的 xlsx"""
    if not os.path.isdir(folder):
        return None, None
    files = glob.glob(os.path.join(folder, "*.xlsx"))
    if not files:
        return None, None
    # 抓檔名中的 8 位數日期（YYYYMMDD），日期最大的優先
    def date_key(p):
        m = re.search(r"(\d{8})", os.path.basename(p))
        if m:
            try:
                return datetime.strptime(m.group(1), "%Y%m%d")
            except ValueError:
                pass
        return datetime.fromtimestamp(os.path.getmtime(p))
    files.sort(key=date_key, reverse=True)
    latest = files[0]
    m = re.search(r"(\d{8})", os.path.basename(latest))
    file_date = m.group(1) if m else None
    return latest, file_date


def _parse_shares(val):
    """解析張數欄位：'20.3 張' / '110股' / '3' → float (張數)"""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # 移除單位
    s_no_unit = re.sub(r"[張股,\s]+", "", s)
    try:
        n = float(s_no_unit)
    except ValueError:
        return None
    # 判讀單位：含「股」且不含「張」→ 股數，要除以 1000
    if "股" in s and "張" not in s:
        return round(n / 1000, 4)
    return n


def _is_etf(sid, name=""):
    """ETF 判讀：代碼 00 開頭"""
    return str(sid).startswith("00")


def load_holdings_from_xlsx(path):
    """讀 xlsx，回傳 list of (sid, name, shares, cost, is_etf)"""
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return None
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return None

    # 找標頭列
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        joined = "".join(str(c) for c in row if c is not None)
        if "股票代號" in joined or "代號" in joined or "代碼" in joined:
            header_idx = i
            break
    header = rows[header_idx]
    # 對應欄位 index
    col = {"sid": None, "name": None, "shares": None, "cost": None,
           "div_cash": None, "div_stock": None, "div_ex_date": None}
    for j, h in enumerate(header):
        if h is None: continue
        s = str(h)
        if "代號" in s or "代碼" in s:        col["sid"] = j
        elif "公司" in s or "名稱" in s:       col["name"] = j
        elif "持股" in s or "張數" in s or "數量" in s: col["shares"] = j
        elif "成本" in s:                      col["cost"] = j
        elif "現金股利" in s:                  col["div_cash"]    = j
        elif "股票股利" in s:                  col["div_stock"]   = j
        elif "除息日" in s or "除權日" in s or "除權息日" in s: col["div_ex_date"] = j
    if col["sid"] is None:
        return None

    holdings = []
    for row in rows[header_idx + 1:]:
        if not row or row[col["sid"]] is None:
            continue
        sid_raw = row[col["sid"]]
        # 整數的 sid 會被 openpyxl 讀成 int，要轉字串；浮點 6279.0 也要轉
        if isinstance(sid_raw, int):
            sid = str(sid_raw)
        elif isinstance(sid_raw, float):
            sid = str(int(sid_raw)) if sid_raw.is_integer() else str(sid_raw)
        else:
            sid = str(sid_raw).strip()
            # 字串如 "6279.0" 也轉成 "6279"（但 "00640L" 保留）
            if re.match(r"^\d+\.0$", sid):
                sid = sid[:-2]
        # 跳過非股票代碼列（例如「股市還有現金」）
        if not re.match(r"^[0-9]{4,6}[A-Za-z]?$", sid):
            continue

        name = str(row[col["name"]]).strip() if col["name"] is not None and row[col["name"]] is not None else sid
        shares = _parse_shares(row[col["shares"]] if col["shares"] is not None else None)
        if shares is None or shares <= 0:
            continue
        cost = None
        if col["cost"] is not None and row[col["cost"]] is not None:
            try:
                c = float(str(row[col["cost"]]).replace(",", "").strip())
                if c > 0:
                    cost = c
            except (ValueError, TypeError):
                pass
        is_etf = _is_etf(sid, name)
        holdings.append((sid, name, shares, cost, is_etf))
    return holdings


def get_holdings():
    """主入口：回傳當前最新的持股清單。
    優先順序：xlsx > 內建預設
    回傳 (holdings, source_file_or_None, file_date_or_None)
    """
    path, file_date = find_latest_holdings_xlsx()
    if path:
        holdings = load_holdings_from_xlsx(path)
        if holdings:
            return holdings, path, file_date
    return DEFAULT_HOLDINGS, None, None


# ── 非持股資產（現金 / 保險 / 勞退）──────────────────

def load_manual_dividends_from_xlsx(path):
    """讀 xlsx「2026 現金股利 / 股票股利 / 除息日」三欄，回傳：
    { sid: {'cash': 元/股, 'stock': 元/股, 'ex_date': 'YYYY-MM-DD' or None } }
    沒填或無欄位則該 sid 不在 dict 內。"""
    out = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return out
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2: return out
    # 找標頭與欄位 index
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        joined = "".join(str(c) for c in row if c is not None)
        if "股票代號" in joined or "代號" in joined or "代碼" in joined:
            header_idx = i; break
    header = rows[header_idx]
    col = {"sid": None, "cash": None, "stock": None, "ex_date": None}
    for j, h in enumerate(header):
        if h is None: continue
        s = str(h)
        if "代號" in s or "代碼" in s: col["sid"] = j
        elif "現金股利" in s:           col["cash"]  = j
        elif "股票股利" in s:           col["stock"] = j
        elif "除息日" in s or "除權日" in s or "除權息日" in s: col["ex_date"] = j
    if col["sid"] is None: return out
    if col["cash"] is None and col["stock"] is None: return out

    def _num(v):
        if v is None: return 0.0
        try: return float(str(v).replace(",", "").strip())
        except (ValueError, TypeError): return 0.0
    def _date(v):
        if v is None: return None
        if hasattr(v, "strftime"): return v.strftime("%Y-%m-%d")
        s = str(v).strip()
        return s if s and s != "—" else None

    for r in rows[header_idx + 1:]:
        if not r or r[col["sid"]] is None: continue
        sid_raw = r[col["sid"]]
        if isinstance(sid_raw, int):     sid = str(sid_raw)
        elif isinstance(sid_raw, float): sid = str(int(sid_raw)) if sid_raw.is_integer() else str(sid_raw)
        else:                            sid = str(sid_raw).strip()
        if re.match(r"^\d+\.0$", sid):   sid = sid[:-2]
        if not re.match(r"^[0-9]{4,6}[A-Za-z]?$", sid): continue
        cash    = _num(r[col["cash"]])  if col["cash"]    is not None else 0.0
        stock   = _num(r[col["stock"]]) if col["stock"]   is not None else 0.0
        ex_date = _date(r[col["ex_date"]]) if col["ex_date"] is not None else None
        if cash > 0 or stock > 0 or ex_date:
            out[sid] = {"cash": cash, "stock": stock, "ex_date": ex_date}
    return out


def load_other_assets_from_xlsx(path):
    """從 xlsx 抽取現金與保險/勞退類項目，回傳 list of (name, value_or_None, note)"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return []
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    items = []
    seen = set()
    # 關鍵字 → 標準名（保證跟 ASSET_EXTRAS_KEYS 對齊）
    KW = [
        ("股市還有現金", "股市現金"),
        ("生活現金",     "生活現金"),
        ("AIA",          "AIA 充裕未來1"),
        ("保誠雋升",     "保誠雋升"),
        ("勞工退休金",   "勞工退休金"),
        ("勞工保險",     "勞工保險"),
    ]
    # xlsx 上現金欄是 B 欄 (idx=1)
    for r in rows:
        if not r: continue
        # 任何欄出現關鍵字
        joined = " | ".join("" if c is None else str(c) for c in r)
        for kw, std_name in KW:
            if kw in joined and std_name not in seen:
                # 找這列裡的數字當預估值
                v = None
                for c in r[1:]:
                    if isinstance(c, (int, float)) and c > 1000:
                        v = float(c)
                        break
                # 抓備註（找有「年」「保單」「投保」「退休」字眼的字串欄）
                note = ""
                for c in r:
                    if isinstance(c, str) and any(k in c for k in ("年", "保單", "投保", "退休", "繳完", "65")):
                        note = c.strip(); break
                items.append({"name": std_name, "value": v, "note": note})
                seen.add(std_name)
    return items


def load_holdings_and_assets():
    """同時讀持股與其他資產，回傳 (holdings, others, source_path, file_date)"""
    path, file_date = find_latest_holdings_xlsx()
    holdings = None
    others = []
    if path:
        holdings = load_holdings_from_xlsx(path)
        others = load_other_assets_from_xlsx(path)
    if not holdings:
        holdings = DEFAULT_HOLDINGS
    return holdings, others, path, file_date


if __name__ == "__main__":
    h, p, d = get_holdings()
    print(f"來源：{p or '內建預設'}")
    if d: print(f"日期：{d}")
    print(f"共 {len(h)} 支：")
    for row in h:
        print(f"  {row[0]:8s} {row[1]:>10s}　{row[2]}張　成本 {row[3]}　ETF={row[4]}")
