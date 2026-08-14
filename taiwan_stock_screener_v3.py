import requests, pandas as pd, time, json, os, pickle, urllib3, openpyxl
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
SESSION = requests.Session()
SESSION.verify = False
SESSION.headers.update({"User-Agent": "Mozilla/5.0", "Referer": "https://www.twse.com.tw/"})

BASE_DIR     = (os.environ.get("STOCK_BASE_DIR") or os.path.expanduser("~/Desktop/Stock Selection Strategy"))
CACHE_DIR    = os.path.join(BASE_DIR, "cache")
OUTPUT_DIR   = BASE_DIR
HISTORY_FILE = os.path.join(CACHE_DIR, "top5_history.json")
os.makedirs(CACHE_DIR, exist_ok=True)
PRICE_CACHE  = os.path.join(CACHE_DIR, "price_data.pkl")
INDEX_CACHE  = os.path.join(CACHE_DIR, "index_data.pkl")
META_FILE    = os.path.join(CACHE_DIR, "meta.json")
HISTORY_DAYS = 200

CFG = {
    "vol_multiple_normal": 2.2, "vol_multiple_gapup": 1.8,
    "bias_limit_initial": 0.08, "bias_limit_overheated": 0.15,
    "pre_break_volume_limit": 1.2, "body_ratio_min": 0.5,
    "upper_shadow_ratio_max": 0.3, "continuation_vol_multiple": 1.5,
    "overheated_vol_multiple": 4.0, "top_warn_vol_multiple": 3.0,
    "score_breakout": 75, "score_watch": 60,
    "liquidity_threshold": 100_000_000, "rs20_threshold": 0.0,
}

def roc_to_date(s):
    p = s.strip().split("/")
    return str(int(p[0])+1911)+"-"+p[1].zfill(2)+"-"+p[2].zfill(2)

def month_range(s, e):
    cur = datetime.strptime(s, "%Y-%m-%d").replace(day=1)
    end = datetime.strptime(e, "%Y-%m-%d")
    while cur <= end:
        yield cur
        cur = cur.replace(month=cur.month+1) if cur.month < 12 else cur.replace(year=cur.year+1, month=1)

def fetch_twse_stock(sid, sd, ed):
    rows = []
    for cur in month_range(sd, ed):
        try:
            r = SESSION.get("https://www.twse.com.tw/rwd/zh/afterTrading/STOCK_DAY",
                params={"date": cur.strftime("%Y%m%d"), "stockNo": sid, "response": "json"}, timeout=15)
            d = r.json()
            if d.get("stat") == "OK":
                for row in d.get("data", []):
                    try:
                        rows.append({"date": roc_to_date(row[0]),
                            "open": float(row[3].replace(",","")), "high": float(row[4].replace(",","")),
                            "low": float(row[5].replace(",","")), "close": float(row[6].replace(",","")),
                            "volume": float(row[1].replace(",","")), "amount": float(row[2].replace(",",""))})
                    except: continue
        except: pass
        time.sleep(1)
    if not rows: return pd.DataFrame()
    df = pd.DataFrame(rows)
    return df[(df["date"]>=sd)&(df["date"]<=ed)].sort_values("date").reset_index(drop=True)

def fetch_tpex_stock(sid, sd, ed):
    # 改用 Yahoo Finance（上櫃API已失效）
    try:
        from datetime import datetime
        r = SESSION.get(f"https://query1.finance.yahoo.com/v8/finance/chart/{sid}.TW",
            params={"interval":"1d","range":"8mo"}, timeout=15,
            headers={"User-Agent":"Mozilla/5.0"})
        if r.status_code != 200: return pd.DataFrame()
        d = r.json()
        result = d.get("chart",{}).get("result")
        if not result: return pd.DataFrame()
        timestamps = result[0]["timestamp"]
        ohlcv = result[0]["indicators"]["quote"][0]
        df = pd.DataFrame({
            "date": [datetime.fromtimestamp(t).strftime("%Y-%m-%d") for t in timestamps],
            "open": ohlcv["open"], "high": ohlcv["high"],
            "low": ohlcv["low"], "close": ohlcv["close"],
            "volume": ohlcv["volume"]
        }).dropna()
        time.sleep(1)
        return df[(df["date"]>=sd)&(df["date"]<=ed)].sort_values("date").reset_index(drop=True)
    except:
        return pd.DataFrame()

def fetch_index(sd, ed):
    rows = []
    for cur in month_range(sd, ed):
        try:
            r = SESSION.get("https://www.twse.com.tw/exchangeReport/FMTQIK",
                params={"date": cur.strftime("%Y%m%d"), "response": "json"}, timeout=15)
            d = r.json()
            if d.get("stat") == "OK":
                for row in d.get("data", []):
                    try: rows.append({"date": roc_to_date(row[0]), "close": float(row[4].replace(",",""))})
                    except: continue
        except: pass
        time.sleep(1)
    if not rows: return pd.DataFrame()
    df = pd.DataFrame(rows)
    return df.sort_values("date").drop_duplicates("date").reset_index(drop=True)

def get_stock_list():
    # 快取股票清單，每天只抓一次
    cache_file = os.path.join(CACHE_DIR, "stock_list_cache.json")
    today = datetime.today().strftime("%Y-%m-%d")
    if os.path.exists(cache_file):
        try:
            with open(cache_file) as f: cached = json.load(f)
            if cached.get("date") == today:
                sd = cached["data"]
                print("  上市+上櫃（快取）："+str(len(sd))+" 支")
                return sd
        except: pass
    sd = {}
    try:
        r = SESSION.get("https://openapi.twse.com.tw/v1/opendata/t187ap03_L", timeout=8)
        for item in r.json():
            s = item.get("公司代號",""); n = item.get("公司簡稱","")
            if s.isdigit() and len(s)==4: sd[s] = {"name": n, "market": "twse"}
        print("  上市："+str(len(sd))+" 支")
    except Exception as e: print("上市失敗："+str(e))
    tc = 0
    try:
        r = SESSION.get("https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes", timeout=8)
        for item in r.json():
            s = str(item.get("SecuritiesCompanyCode","")).strip(); n = item.get("CompanyAbbreviation","")
            if s.isdigit() and len(s)==4 and s not in sd:
                sd[s] = {"name": n, "market": "tpex"}; tc += 1
        print("  上櫃："+str(tc)+" 支")
    except Exception as e: print("上櫃失敗："+str(e))
    if not sd: raise RuntimeError("無法取得股票清單")
    print("合計："+str(len(sd))+" 支")
    # 存快取
    with open(cache_file,"w") as f: json.dump({"date":today,"data":sd},f)
    return sd

def load_cache():
    pc = {}; idx = {}; meta = {"last_update": None, "stock_list": []}
    if os.path.exists(PRICE_CACHE):
        with open(PRICE_CACHE,"rb") as f: pc = pickle.load(f)
    if os.path.exists(INDEX_CACHE):
        with open(INDEX_CACHE,"rb") as f: idx = pickle.load(f)
    if os.path.exists(META_FILE):
        with open(META_FILE,"r") as f: meta = json.load(f)
    return pc, idx, meta

def save_cache(pc, idx, meta):
    with open(PRICE_CACHE,"wb") as f: pickle.dump(pc, f)
    with open(INDEX_CACHE,"wb") as f: pickle.dump(idx, f)
    with open(META_FILE,"w") as f: json.dump(meta, f, ensure_ascii=False)

def full_init(sd):
    ed = (datetime.today()-timedelta(days=1)).strftime("%Y-%m-%d")
    s  = (datetime.today()-timedelta(days=HISTORY_DAYS)).strftime("%Y-%m-%d")
    sl = list(sd.keys()); pc, idx, _ = load_cache()
    print("\n下載歷史資料（已有"+str(len(pc))+"/"+str(len(sl))+" 支，斷點續傳）")
    print("="*55)
    if "TAIEX" not in idx:
        print("  下載加權指數...")
        idf = fetch_index(s, ed)
        if not idf.empty: idx["TAIEX"] = idf; print("  指數："+str(len(idf))+" 筆")
        time.sleep(1)
    done = saved = 0
    for sid in sl:
        done += 1
        if sid in pc:
            print("  "+str(done)+"/"+str(len(sl))+" 快取："+str(len(pc))+" 支", end="\r"); continue
        market = sd[sid]["market"]
        df = fetch_twse_stock(sid,s,ed) if market=="twse" else fetch_tpex_stock(sid,s,ed)
        if not df.empty and len(df)>=10: pc[sid] = df
        saved += 1
        print("  "+str(done)+"/"+str(len(sl))+" ("+str(int(done/len(sl)*100))+"%) - "+sid+" | 快取："+str(len(pc))+" 支", end="\r")
        if saved % 50 == 0:
            save_cache(pc, idx, {"last_update": None, "start_date": s, "stock_list": sl})
            print("\n  💾 自動存檔（"+str(len(pc))+" 支）...")
    meta = {"last_update": ed, "start_date": s, "stock_list": sl}
    save_cache(pc, idx, meta)
    print("\n✅ 下載完成！共"+str(len(pc))+" 支")
    return pc, idx, meta

def incremental_update(pc, idx, meta, sd):
    """
    增量更新：先看每股 cache 自身的最新日期，已 up-to-date 就跳過。
    解決 1966 支股 × TWSE API 過慢造成 900s 超時的問題。
    daily_yahoo_update.py 跑過後，這個函數通常什麼都不做（直接 return）。
    """
    if datetime.now().hour >= 14:
        yd = datetime.today().strftime("%Y-%m-%d")
    else:
        yd = (datetime.today()-timedelta(days=1)).strftime("%Y-%m-%d")

    # ★ 新邏輯：先檢查 cache 內個別股票最新日期
    cached_sids = [s for s in pc.keys() if s in sd]
    stale_sids = []   # 真正需要 update 的股
    for sid in cached_sids:
        df = pc.get(sid)
        if df is None or df.empty:
            stale_sids.append(sid); continue
        try:
            last = str(df["date"].iloc[-1])
        except Exception:
            stale_sids.append(sid); continue
        if last < yd:
            stale_sids.append(sid)

    if not stale_sids:
        print(f"✅ 資料已是最新（cache 內 {len(cached_sids)} 支皆達 {yd}）")
        meta["last_update"] = yd
        save_cache(pc, idx, meta)
        return pc, idx

    # 若超過 50% 股票需更新，停止用 TWSE API（會超時）— 改提示用 daily_yahoo_update
    if len(stale_sids) > len(cached_sids) * 0.5:
        print(f"⚠️ {len(stale_sids)}/{len(cached_sids)} 支需更新 — TWSE API 太慢，建議先跑 daily_yahoo_update.py")
        print(f"   仍嘗試以 TWSE API 增量更新前 100 支，其餘留待下次...")
        stale_sids = stale_sids[:100]

    ld_str = meta.get("last_update", (datetime.today()-timedelta(days=7)).strftime("%Y-%m-%d"))
    s_date = (datetime.strptime(ld_str,"%Y-%m-%d")+timedelta(days=1)).strftime("%Y-%m-%d") if ld_str < yd else yd
    print(f"\n增量更新 {len(stale_sids)} 支：{s_date} ~ {yd}")

    # 大盤指數也補
    ni = fetch_index(s_date, yd)
    if not ni.empty:
        old = idx.get("TAIEX", pd.DataFrame())
        idx["TAIEX"] = pd.concat([old,ni]).drop_duplicates("date").sort_values("date").reset_index(drop=True)

    cutoff = (datetime.today()-timedelta(days=HISTORY_DAYS)).strftime("%Y-%m-%d")
    done = 0
    for sid in stale_sids:
        done += 1
        info = sd[sid]
        ndf = fetch_twse_stock(sid,s_date,yd) if info["market"]=="twse" else fetch_tpex_stock(sid,s_date,yd)
        if not ndf.empty:
            old = pc.get(sid, pd.DataFrame())
            merged = pd.concat([old,ndf]).drop_duplicates("date").sort_values("date")
            pc[sid] = merged[merged["date"]>=cutoff].reset_index(drop=True)
        print(f"  更新：{done}/{len(stale_sids)} ({sid})", end="\r", flush=True)
    meta["last_update"] = yd
    save_cache(pc, idx, meta)
    print(f"\n✅ 增量更新完成（{done} 支）")
    return pc, idx

def risk_label(bias_str):
    try:
        bias = float(bias_str.replace("%",""))
        if bias >= 20: return "🔴 高風險"
        elif bias >= 10: return "🟡 注意"
        else: return "🟢 正常"
    except: return "-"

def risk_color(bias_str):
    try:
        bias = float(bias_str.replace("%",""))
        if bias >= 20: return "C00000"
        elif bias >= 10: return "E36C09"
        else: return "217346"
    except: return "333333"

def evaluate_stock(sid, df, idf):
    if df is None or df.empty or len(df) < 22: return None
    for c in ["open","high","low","close","volume"]:
        if c in df.columns: df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df.dropna(subset=["close","volume"]).reset_index(drop=True)
    if len(df) < 22: return None
    t = len(df)-1
    O,H,L,C,V = df["open"].iloc[t],df["high"].iloc[t],df["low"].iloc[t],df["close"].iloc[t],df["volume"].iloc[t]
    MA5  = df["close"].iloc[t-4:t+1].mean()
    MA10 = df["close"].iloc[t-9:t+1].mean()
    MA20 = df["close"].iloc[t-19:t+1].mean()
    VMA20 = df["volume"].iloc[t-19:t+1].mean()
    H20 = df["high"].iloc[max(0,t-20):t].max()
    Rng = H-L; B20 = C/MA20-1 if MA20>0 else 0
    VB = V/VMA20 if VMA20>0 else 0
    BP = C/H20-1 if H20>0 else 0
    BR = abs(C-O)/Rng if Rng>0 else 0
    US = (H-max(O,C))/Rng if Rng>0 else 1
    GU = L > df["high"].iloc[t-1] if t>0 else False
    A20 = (df["close"]*df["volume"]).iloc[t-19:t+1].mean()
    pv = df["volume"].iloc[max(0,t-10):t]
    DR = (pv < CFG["pre_break_volume_limit"]*VMA20).sum() / max(1,len(pv))
    RS = 0
    if idf is not None and not idf.empty and t >= 20:
        s20 = C/df["close"].iloc[t-20]-1 if df["close"].iloc[t-20]>0 else 0
        ic = idf["close"].dropna()
        if len(ic) >= 21: RS = s20-(ic.iloc[-1]/ic.iloc[-21]-1)
    RV = CFG["vol_multiple_gapup"] if GU else CFG["vol_multiple_normal"]
    G1=VB>=RV; G2=C>H20; G3=B20<CFG["bias_limit_initial"]
    G4 = df["volume"].iloc[max(0,t-5):t].max() < CFG["pre_break_volume_limit"]*VMA20 if t>=5 else True
    G5 = BR>CFG["body_ratio_min"] and US<CFG["upper_shadow_ratio_max"]
    G6=A20>=CFG["liquidity_threshold"]; G7=RS>CFG["rs20_threshold"]
    Gate = G1 and G2 and G3 and G4 and G5 and G6 and G7
    S1=20*min(1,VB/3); S2=20*min(1,max(0,BP)/0.05); S3=15*DR
    S4 = 10 if B20<=0.08 else max(0,10*(1-(B20-0.08)/0.07))
    S5 = 10*((1 if MA5>MA10 else 0)+(1 if MA10>MA20 else 0))/2
    S6 = 15*min(1,BR/0.7)*(1-min(1,US/0.5))
    S7 = 10 if RS>=0.10 else max(0,10*RS/0.10)
    Sc = S1+S2+S3+S4+S5+S6+S7

    # ─── 飆股進階訊號（基於 5/4→5/15 回測：高分=末升段、中分=起漲）─────
    # 設計依據：
    #   1) 高評分 + 高乖離 + 高 RSI = 已漲多，命中率反而低 → 扣分
    #   2) 窄幅打底（近 60 日 std/mean ≤ 8%）+ 突破箱頂 + 量爆 = 真起漲 → 加分
    #   3) 距年線（MA240）> 30% = 末升段 → 扣分
    starter_bonus = 0
    overext_penalty = 0
    starter_flag = False
    if t >= 60:
        c60 = df["close"].iloc[t-59:t+1]
        consol = (c60.std() / c60.mean()) if c60.mean() > 0 else 1
        box_top60 = df["high"].iloc[t-59:t].max() if t >= 60 else float("inf")
        breakout_box = C > box_top60
        avg20v_prev = df["volume"].iloc[t-20:t].mean()
        vol_2x = V >= avg20v_prev * 2
        is_red = C > O
        if consol <= 0.08 and breakout_box and vol_2x and is_red:
            starter_bonus = 15   # 真起漲三件套：打底窄幅 + 突破箱頂 + 量爆
            starter_flag = True
        elif consol <= 0.10 and breakout_box:
            starter_bonus = 8

        gain60 = (C / df["close"].iloc[t-59] - 1) if df["close"].iloc[t-59] > 0 else 0
        # 5/4→5/15 回測：聯策 RSI87 / 6658 仍漲 35%，過熱扣分易誤殺主升段飆股
        # 因此只扣「極端過熱」（>80% 漲幅、>RSI 85），中等過熱不扣
        if gain60 > 0.80:    overext_penalty += 15
        elif gain60 > 0.50:  overext_penalty += 6

    if t >= 240:
        ma240 = df["close"].iloc[t-239:t+1].mean()
        gap240 = (C / ma240 - 1) if ma240 > 0 else 0
        if gap240 > 1.00:    overext_penalty += 10
        elif gap240 > 0.60:  overext_penalty += 4

    # RSI(14)
    delta_arr = df["close"].diff()
    up14 = delta_arr.clip(lower=0).iloc[t-13:t+1].mean()
    dn14 = (-delta_arr.clip(upper=0)).iloc[t-13:t+1].mean()
    rsi14 = 100 - 100/(1 + (up14/dn14)) if dn14 > 0 else 50
    if rsi14 > 90:        overext_penalty += 5

    Sc = max(0, Sc + starter_bonus - overext_penalty)

    if Gate and Sc>=CFG["score_breakout"]:   Sig="INITIAL_BREAKOUT"
    elif Gate and Sc>=CFG["score_watch"]:    Sig="BREAKOUT_WATCH"
    elif C>MA5 and B20>=0.08 and VB>CFG["continuation_vol_multiple"]: Sig="CONTINUATION"
    # 過熱需同時滿足：乖離率過高 AND 量比過大（避免剛起漲被誤判）
    elif B20>CFG["bias_limit_overheated"] and VB>CFG["overheated_vol_multiple"]: Sig="OVERHEATED"
    # 只有乖離率高但量比正常 → 續漲
    elif B20>CFG["bias_limit_overheated"] and C>MA5: Sig="CONTINUATION"
    # 只有量比大但乖離率低 → 起漲觀察
    elif VB>CFG["overheated_vol_multiple"] and B20<=CFG["bias_limit_overheated"]: Sig="BREAKOUT_WATCH"
    else: Sig="NONE"
    pM = df["close"].iloc[t-10:t].mean() if t>=10 else MA10
    Top = (VB>CFG["top_warn_vol_multiple"]*VMA20 and C<O) or ((C>H20 and US>0.4) and (C<MA10 and MA10<pM))
    # V34/V42 命中：即使 Sig=NONE 也納入（這些是高機率飆股，不能漏）
    # 必須先算出 v34_hit（往下幾行）才能決定是否 return None，
    # 所以先暫存決定，到下面 v34_hit 算完後再判斷。
    bias_str = str(round(B20*100,1))+"%"
    # ─── V42_R6 冠軍策略（2 年全市場回測 + 族群過濾驗證）──
    # 56 組合矩陣回測冠軍：V42_R6 + S6 Top 7 子族群
    # 結果：+10% 命中率 32.4%（vs 原 V42 22.2%，提升 +10.2pt / 46%）
    #       平均報酬 +6.62%、虧損率 36.6%（最低）、樣本 71 統計最可信
    # 條件（10 條全部成立 → 🌟 高機率飆股）：
    #   1) MA5 > MA20             （短線多頭）
    #   2) MA20 > MA60            （中期多頭排列）
    #   3) MA20 斜率 > 1.5%       （月線上揚，從 2.5% 放寬）
    #   4) 量比 ≥ 1.0             （量價配合，從 1.3 放寬）
    #   5) 不是高檔變盤線
    #   6) 不是連跌轉空
    #   7) RSI14 < 80
    #   8) K 棒實體 > 40%
    #   9) RS20 > 5%              （強於大盤，從 10% 放寬）
    #  10) 收盤在 MA20 × 1.02 ~ 1.30  （從 1.10 放寬到 1.30，可進主升段）
    # ⚠️ 另需「至少一個歸屬族群在子族群 Top 7」雙重確認（在 build_flash_picks 過濾）
    MA60_local = df["close"].iloc[t-59:t+1].mean() if t >= 59 else 0
    ma20_5ago_local = df["close"].iloc[t-24:t-4].mean() if t >= 24 else 0
    ma20_slope_local = (MA20 - ma20_5ago_local) / ma20_5ago_local if ma20_5ago_local > 0 else 0

    # 變盤線/連跌判定（用 advanced_signals）
    is_doji_top_local = False
    is_cons_down_local = False
    try:
        from advanced_signals import doji_reversal as _dr, consecutive_down_after_up as _cd
        _drr = _dr(df); _cdr = _cd(df)
        is_doji_top_local = (_drr.get("signal") == "top")
        is_cons_down_local = _cdr.get("signal", False)
    except Exception:
        pass

    v34_hit = (
        MA5 > MA20
        and MA20 > MA60_local
        and ma20_slope_local > 0.015     # V42_R6 放寬：1.5%
        and VB >= 1.0                     # V42_R6 放寬：量比 1.0
        and not is_doji_top_local
        and not is_cons_down_local
        and rsi14 < 80
        and BR > 0.4
        and RS > 0.05                     # V42_R6 放寬：RS 5%
        and C > MA20 * 1.02
        and C < MA20 * 1.30               # V42_R6 放寬：上限 1.30
    )

    # V42 命中時若 Sig=NONE，覆寫為 V42_FLASH（高機率飆股獨立訊號）
    if v34_hit and Sig == "NONE":
        Sig = "V42_FLASH"

    # 真正 return None 的條件：Sig=NONE 且非 V42 命中且非高檔反轉
    if Sig == "NONE" and not Top:
        return None
    return {
        "股票代碼": sid, "訊號": Sig, "收盤價": round(C,2),
        "評分": round(Sc,1), "量比": round(VB,2),
        "乖離率": bias_str, "突破幅度": str(round(BP*100,1))+"%",
        "K棒品質": round(BR,2), "RS20": str(round(RS*100,1))+"%",
        "高檔反轉": "是" if Top else "否", "Gate": "Y" if Gate else "N",
        "今日收盤": round(C,2), "今日量": round(V,0), "VMA20": round(VMA20,0),
        "突破點": round(H20,2), "風險警示": risk_label(bias_str),
        "MA5": round(MA5,2), "MA10": round(MA10,2), "MA20": round(MA20,2),
        "起漲點": "🚀" if starter_flag else "",
        "過熱扣分": round(overext_penalty, 1),
        "RSI14": round(rsi14, 1),
        "飆股命中": "🌟" if v34_hit else "",   # V34 冠軍策略命中
        "MA60": round(MA60_local, 2),
        "月線斜率": round(ma20_slope_local * 100, 2),  # %
    }

def load_top5_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE,"r",encoding="utf-8") as f: return json.load(f)
    return {}

def save_top5_history(history):
    with open(HISTORY_FILE,"w",encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

def diagnose_stock(sid, pc, yesterday_data):
    df = pc.get(sid)
    if df is None or df.empty or len(df) < 3: return "❓", "資料不足，無法判斷"
    df = df.copy()
    for c in ["open","high","low","close","volume"]:
        if c in df.columns: df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df.dropna(subset=["close","volume"]).reset_index(drop=True)
    if len(df) < 2: return "❓", "資料不足"
    today = df.iloc[-1]; prev = df.iloc[-2]
    C_today = today["close"]; C_prev = prev["close"]
    V_today = today["volume"]; V_prev = prev["volume"]
    VMA20 = df["volume"].iloc[max(0,len(df)-20):len(df)].mean()
    H20_prev = yesterday_data.get("突破點", C_prev)
    chg_pct = (C_today - C_prev) / C_prev * 100 if C_prev > 0 else 0
    vol_ratio = V_today / V_prev if V_prev > 0 else 1
    if C_today > C_prev and vol_ratio >= 1.2:
        return "✅ 延續漲勢", "今日漲幅 +"+str(round(chg_pct,1))+"% ，量增 "+str(round(vol_ratio,1))+"x，多頭延續中"
    elif C_today > C_prev and vol_ratio < 0.8:
        return "⚠️ 量縮上漲", "今日漲幅 +"+str(round(chg_pct,1))+"%，但量縮至 "+str(round(vol_ratio,1))+"x，動能不足，需觀察"
    elif C_today > C_prev:
        return "🔶 小幅上漲", "今日漲幅 +"+str(round(chg_pct,1))+"%，量能持平，繼續觀察"
    elif C_today < H20_prev:
        return "❌ 假突破", "今日收盤 "+str(C_today)+" 已跌回突破點 "+str(H20_prev)+" 以下，起漲訊號失效"
    elif C_today < C_prev and V_today > VMA20 * 1.5:
        return "🔴 爆量下跌", "今日跌幅 "+str(round(chg_pct,1))+"%，爆量 "+str(round(vol_ratio,1))+"x，主力出貨訊號"
    elif C_today < C_prev and vol_ratio < 0.8:
        return "🔶 量縮回測", "今日跌幅 "+str(round(chg_pct,1))+"%，量縮正常回測，突破點 "+str(H20_prev)+" 仍需守住"
    elif C_today < C_prev:
        return "⚠️ 小幅回跌", "今日跌幅 "+str(round(chg_pct,1))+"%，未跌破突破點，持續觀察"
    else:
        return "➡️ 橫盤整理", "今日漲跌幅 "+str(round(chg_pct,1))+"%，橫盤整理中"

# ── HTML K線圖報表 ────────────────────────────────────

def generate_html(results, sd, pc, output_path):
    """產生互動式 HTML K線圖報表"""
    # ★ 過濾空字串名稱（上櫃股在 stock_list_cache 內常為 ''）
    names  = {s: i["name"] for s, i in sd.items()
              if i.get("name") and str(i["name"]).strip()}
    # Fallback：從 sector_analyzer.fetch_all_industries 補齊
    try:
        from sector_analyzer import fetch_all_industries
        ind = fetch_all_industries()
        for sid, info in ind.items():
            existing = names.get(sid, "")
            if (not existing or not str(existing).strip()) and info.get("name"):
                names[sid] = info["name"]
    except Exception:
        pass
    labels = {"INITIAL_BREAKOUT":"🚀 初始起漲","BREAKOUT_WATCH":"👀 起漲觀察",
              "CONTINUATION":"📈 續漲","OVERHEATED":"🔥 過熱警示"}
    sig_colors_hex = {"INITIAL_BREAKOUT":"#1F497D","BREAKOUT_WATCH":"#2E75B6",
                      "CONTINUATION":"#217346","OVERHEATED":"#C00000"}
    today_str = datetime.today().strftime("%Y/%m/%d")

    # 產生每支股票的K線圖 JSON 資料
    stock_charts = []
    for r in results:
        sid = r["股票代碼"]
        df = pc.get(sid)
        if df is None or df.empty: continue
        df = df.copy()
        for c in ["open","high","low","close","volume"]:
            if c in df.columns: df[c] = pd.to_numeric(df[c], errors="coerce")
        df = df.dropna(subset=["open","high","low","close","volume"]).tail(60)  # 最近60天

        # 計算均線
        df["ma5"]  = df["close"].rolling(5).mean().round(2)
        df["ma10"] = df["close"].rolling(10).mean().round(2)
        df["ma20"] = df["close"].rolling(20).mean().round(2)

        dates  = df["date"].tolist()
        opens  = df["open"].tolist()
        highs  = df["high"].tolist()
        lows   = df["low"].tolist()
        closes = df["close"].tolist()
        vols   = df["volume"].tolist()
        ma5    = df["ma5"].tolist()
        ma10   = df["ma10"].tolist()
        ma20   = df["ma20"].tolist()

        stock_charts.append({
            "sid": sid,
            "name": names.get(sid, sid),
            "signal": r["訊號"],
            "signal_label": labels.get(r["訊號"], r["訊號"]),
            "score": r["評分"],
            "bias": r["乖離率"],
            "risk": r["風險警示"],
            "rs20": r["RS20"],
            "vol_ratio": r["量比"],
            "dates": dates,
            "opens": opens,
            "highs": highs,
            "lows": lows,
            "closes": closes,
            "volumes": vols,
            "ma5": ma5,
            "ma10": ma10,
            "ma20": ma20,
            "signal_color": sig_colors_hex.get(r["訊號"], "#333333"),
        })

    charts_json = json.dumps(stock_charts, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>台股飆股選股日報 {today_str}</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: Arial, sans-serif; background: #f0f2f5; color: #333; }}
.header {{ background: linear-gradient(135deg, #1F497D, #2E75B6); color: white; padding: 20px 30px; }}
.header h1 {{ font-size: 24px; margin-bottom: 5px; }}
.header p {{ font-size: 14px; opacity: 0.85; }}
.summary {{ display: flex; gap: 15px; padding: 15px 30px; background: white; border-bottom: 1px solid #ddd; flex-wrap: wrap; }}
.summary-item {{ padding: 8px 16px; border-radius: 20px; font-size: 13px; font-weight: bold; }}
.sig-initial {{ background: #1F497D; color: white; }}
.sig-watch {{ background: #2E75B6; color: white; }}
.sig-cont {{ background: #217346; color: white; }}
.sig-hot {{ background: #C00000; color: white; }}
.filter-bar {{ padding: 12px 30px; background: #fff; border-bottom: 1px solid #eee; display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }}
.filter-bar label {{ font-size: 13px; color: #555; }}
.filter-btn {{ padding: 5px 14px; border: 1px solid #ddd; border-radius: 15px; cursor: pointer; font-size: 12px; background: white; }}
.filter-btn.active {{ background: #1F497D; color: white; border-color: #1F497D; }}
.search-box {{ padding: 5px 12px; border: 1px solid #ddd; border-radius: 15px; font-size: 13px; width: 180px; }}
.grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(580px, 1fr)); gap: 20px; padding: 20px 30px; }}
.card {{ background: white; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); overflow: hidden; }}
.card-header {{ padding: 12px 16px; display: flex; justify-content: space-between; align-items: center; }}
.card-title {{ font-size: 16px; font-weight: bold; }}
.card-signal {{ padding: 3px 10px; border-radius: 12px; font-size: 12px; font-weight: bold; color: white; }}
.card-meta {{ display: flex; gap: 12px; padding: 0 16px 10px; font-size: 12px; color: #555; flex-wrap: wrap; }}
.meta-item {{ display: flex; gap: 4px; }}
.meta-label {{ color: #999; }}
.meta-value {{ font-weight: bold; }}
.risk-high {{ color: #C00000; }}
.risk-mid {{ color: #E36C09; }}
.risk-low {{ color: #217346; }}
.chart-container {{ padding: 0 8px 8px; }}
canvas {{ width: 100% !important; }}
.no-results {{ text-align: center; padding: 60px; color: #999; font-size: 16px; }}
</style>
</head>
<body>
<div class="header">
  <h1>📈 台股飆股選股日報</h1>
  <p>報告日期：{today_str}　共 {len(stock_charts)} 支有訊號股票</p>
</div>
<div class="summary" id="summary"></div>
<div class="filter-bar">
  <label>篩選：</label>
  <button class="filter-btn active" onclick="filterSig('all')">全部</button>
  <button class="filter-btn" onclick="filterSig('INITIAL_BREAKOUT')">🚀 初始起漲</button>
  <button class="filter-btn" onclick="filterSig('BREAKOUT_WATCH')">👀 起漲觀察</button>
  <button class="filter-btn" onclick="filterSig('CONTINUATION')">📈 續漲</button>
  <button class="filter-btn" onclick="filterSig('OVERHEATED')">🔥 過熱</button>
  <input class="search-box" type="text" placeholder="搜尋代碼/名稱..." oninput="searchStock(this.value)">
</div>
<div class="grid" id="grid"></div>

<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script>
const DATA = {charts_json};
const SIG_COLORS = {{
  INITIAL_BREAKOUT: '#1F497D', BREAKOUT_WATCH: '#2E75B6',
  CONTINUATION: '#217346', OVERHEATED: '#C00000'
}};
const SIG_LABELS = {{
  INITIAL_BREAKOUT: '🚀 初始起漲', BREAKOUT_WATCH: '👀 起漲觀察',
  CONTINUATION: '📈 續漲', OVERHEATED: '🔥 過熱警示'
}};

let currentFilter = 'all';
let currentSearch = '';
const charts = {{}};

function riskClass(risk) {{
  if (risk.includes('高')) return 'risk-high';
  if (risk.includes('注意')) return 'risk-mid';
  return 'risk-low';
}}

function renderSummary() {{
  const counts = {{}};
  DATA.forEach(d => counts[d.signal] = (counts[d.signal]||0)+1);
  const el = document.getElementById('summary');
  el.innerHTML = [
    ['INITIAL_BREAKOUT','sig-initial'],['BREAKOUT_WATCH','sig-watch'],
    ['CONTINUATION','sig-cont'],['OVERHEATED','sig-hot']
  ].map(([sig,cls]) => counts[sig] ?
    `<div class="summary-item ${{cls}}">${{SIG_LABELS[sig]}}：${{counts[sig]}} 支</div>` : ''
  ).join('');
}}

function renderChart(stock) {{
  const canvas = document.getElementById('chart-'+stock.sid);
  if (!canvas) return;
  if (charts[stock.sid]) {{ charts[stock.sid].destroy(); }}

  const n = stock.dates.length;
  const candleData = stock.dates.map((d,i) => ({{
    x: i, o: stock.opens[i], h: stock.highs[i], l: stock.lows[i], c: stock.closes[i]
  }}));

  // K線用 bar chart 模擬（上影+下影+實體）
  const barColors = stock.closes.map((c,i) => c >= stock.opens[i] ? 'rgba(220,53,69,0.85)' : 'rgba(40,167,69,0.85)');
  const barBorders = stock.closes.map((c,i) => c >= stock.opens[i] ? '#dc3545' : '#28a745');

  charts[stock.sid] = new Chart(canvas, {{
    type: 'bar',
    data: {{
      labels: stock.dates.map(d => d.slice(5)),
      datasets: [
        {{
          label: '收盤價',
          data: stock.closes,
          backgroundColor: barColors,
          borderColor: barBorders,
          borderWidth: 1,
          yAxisID: 'y',
          order: 2,
        }},
        {{
          label: 'MA5',
          data: stock.ma5,
          type: 'line', borderColor: '#FF9800', borderWidth: 1.5,
          pointRadius: 0, fill: false, yAxisID: 'y', order: 1, tension: 0.3,
        }},
        {{
          label: 'MA10',
          data: stock.ma10,
          type: 'line', borderColor: '#2196F3', borderWidth: 1.5,
          pointRadius: 0, fill: false, yAxisID: 'y', order: 1, tension: 0.3,
        }},
        {{
          label: 'MA20',
          data: stock.ma20,
          type: 'line', borderColor: '#9C27B0', borderWidth: 1.5,
          pointRadius: 0, fill: false, yAxisID: 'y', order: 1, tension: 0.3,
        }},
        {{
          label: '成交量',
          data: stock.volumes.map(v => v/1000),
          backgroundColor: barColors.map(c => c.replace('0.85','0.4')),
          yAxisID: 'y2', order: 3,
        }},
      ]
    }},
    options: {{
      responsive: true,
      interaction: {{ mode: 'index', intersect: false }},
      plugins: {{
        legend: {{ labels: {{ font: {{ size: 10 }}, boxWidth: 20 }} }},
        tooltip: {{
          callbacks: {{
            label: (ctx) => {{
              if (ctx.dataset.label === '成交量') return '量：' + (ctx.raw*1000).toLocaleString();
              return ctx.dataset.label + '：' + ctx.raw;
            }}
          }}
        }}
      }},
      scales: {{
        x: {{ ticks: {{ font: {{ size: 9 }}, maxTicksLimit: 12 }}, grid: {{ display: false }} }},
        y: {{ position: 'left', ticks: {{ font: {{ size: 9 }} }}, grid: {{ color: '#f0f0f0' }} }},
        y2: {{ position: 'right', ticks: {{ font: {{ size: 9 }} }}, grid: {{ display: false }} }},
      }}
    }}
  }});
}}

function renderCards(data) {{
  const grid = document.getElementById('grid');
  if (data.length === 0) {{
    grid.innerHTML = '<div class="no-results">沒有符合條件的股票</div>';
    return;
  }}
  grid.innerHTML = data.map(s => `
    <div class="card" data-sig="${{s.signal}}" data-name="${{s.name}}" data-sid="${{s.sid}}">
      <div class="card-header">
        <div class="card-title">${{s.sid}} ${{s.name}}</div>
        <span class="card-signal" style="background:${{s.signal_color}}">${{s.signal_label}}</span>
      </div>
      <div class="card-meta">
        <div class="meta-item"><span class="meta-label">評分</span><span class="meta-value">${{s.score}}</span></div>
        <div class="meta-item"><span class="meta-label">量比</span><span class="meta-value">${{s.vol_ratio}}x</span></div>
        <div class="meta-item"><span class="meta-label">乖離率</span><span class="meta-value">${{s.bias}}</span></div>
        <div class="meta-item"><span class="meta-label">RS20</span><span class="meta-value">${{s.rs20}}</span></div>
        <div class="meta-item"><span class="meta-label">風險</span><span class="meta-value ${{riskClass(s.risk)}}">${{s.risk}}</span></div>
      </div>
      <div class="chart-container"><canvas id="chart-${{s.sid}}" height="200"></canvas></div>
    </div>
  `).join('');
  // render charts after DOM update
  setTimeout(() => data.forEach(s => renderChart(s)), 50);
}}

function filterSig(sig) {{
  currentFilter = sig;
  document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
  event.target.classList.add('active');
  applyFilter();
}}

function searchStock(val) {{
  currentSearch = val.toLowerCase();
  applyFilter();
}}

function applyFilter() {{
  let filtered = DATA;
  if (currentFilter !== 'all') filtered = filtered.filter(d => d.signal === currentFilter);
  if (currentSearch) filtered = filtered.filter(d =>
    d.sid.includes(currentSearch) || d.name.toLowerCase().includes(currentSearch));
  // destroy old charts
  Object.values(charts).forEach(c => c.destroy());
  Object.keys(charts).forEach(k => delete charts[k]);
  renderCards(filtered);
}}

renderSummary();
renderCards(DATA);
</script>
</body>
</html>"""

    html_path = output_path.replace(".xlsx", ".html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print("✅ HTML K線圖報表已產生："+html_path)
    return html_path

# ── Excel 報表 ────────────────────────────────────────

def make_cell(ws, row, col, value, fill_color=None, font_color="000000",
              bold=False, size=10, align="center", border=None, hyperlink=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", size=size, bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fill_color: cell.fill = PatternFill("solid", fgColor=fill_color)
    if border: cell.border = border
    if hyperlink: cell.hyperlink = hyperlink; cell.font = Font(name="Arial",size=size,color="0563C1",underline="single")
    return cell

def generate_excel(results, sd, pc, output_path):
    order = {"INITIAL_BREAKOUT":0,"BREAKOUT_WATCH":1,"CONTINUATION":2,"OVERHEATED":3}
    results.sort(key=lambda x:(order.get(x["訊號"],9),-x["評分"]))
    names  = {s:i["name"] for s,i in sd.items() if i.get("name")}
    # ★ Fallback：從 sector_analyzer.fetch_all_industries 補齊
    try:
        from sector_analyzer import fetch_all_industries
        ind = fetch_all_industries()
        for sid, info in ind.items():
            if sid not in names and info.get("name"):
                names[sid] = info["name"]
    except Exception:
        pass
    labels = {"INITIAL_BREAKOUT":"🚀 初始起漲","BREAKOUT_WATCH":"👀 起漲觀察",
              "CONTINUATION":"📈 續漲","OVERHEATED":"🔥 過熱警示"}
    sig_colors = {"INITIAL_BREAKOUT":"1F497D","BREAKOUT_WATCH":"2E75B6",
                  "CONTINUATION":"217346","OVERHEATED":"C00000"}
    thin = Side(style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    wb = openpyxl.Workbook()

    # ── Sheet 1：今日選股 ──────────────────────────────
    ws1 = wb.active; ws1.title = "今日選股"
    today_str = datetime.today().strftime("%Y/%m/%d")
    ws1.merge_cells("A1:N1")
    make_cell(ws1,1,1,"📈 台股飆股選股日報　"+today_str,
              fill_color="1F497D", font_color="FFFFFF", bold=True, size=16)
    ws1.row_dimensions[1].height = 32
    ti=len([r for r in results if r["訊號"]=="INITIAL_BREAKOUT"])
    tw=len([r for r in results if r["訊號"]=="BREAKOUT_WATCH"])
    tc=len([r for r in results if r["訊號"]=="CONTINUATION"])
    to=len([r for r in results if r["訊號"]=="OVERHEATED"])
    ws1.merge_cells("A2:N2")
    make_cell(ws1,2,1,
        "🚀 初始起漲："+str(ti)+"　👀 起漲觀察："+str(tw)+"　📈 續漲："+str(tc)+"　🔥 過熱："+str(to)+"　共"+str(len(results))+"支",
        fill_color="EBF3FB", bold=True, size=11)
    ws1.row_dimensions[2].height = 22
    ws1.merge_cells("A3:N3")
    make_cell(ws1,3,1,"風險警示：🟢 乖離<10% 正常　🟡 乖離10~20% 注意追高　🔴 乖離>20% 高風險　　點擊「K線圖」欄可查看詳細圖表",
        fill_color="FFF2CC", font_color="7F6000", size=10, align="left")
    ws1.row_dimensions[3].height = 18

    headers = ["代碼","股票名稱","訊號","評分","量比","乖離率","風險警示","突破幅度","K棒品質","RS20","高檔反轉","Gate","收盤價","K線圖"]
    col_w   = [8,12,12,7,7,8,10,8,8,8,8,6,8,10]
    for i,(h,w) in enumerate(zip(headers,col_w),1):
        make_cell(ws1,4,i,h,fill_color="1F497D",font_color="FFFFFF",bold=True,border=border)
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[4].height = 20

    row_num = 5; prev_sig = ""
    for r in results:
        sig = r["訊號"]
        if sig != prev_sig:
            ws1.merge_cells("A"+str(row_num)+":N"+str(row_num))
            cnt = len([x for x in results if x["訊號"]==sig])
            make_cell(ws1,row_num,1,labels.get(sig,sig)+"　共 "+str(cnt)+" 支",
                fill_color=sig_colors.get(sig,"444444"),font_color="FFFFFF",bold=True,size=11,align="left")
            ws1.row_dimensions[row_num].height = 20; row_num += 1; prev_sig = sig

        bg = "FFFFFF" if row_num%2==0 else "F5F5F5"
        bias_str = r["乖離率"]
        rlabel = r.get("風險警示", risk_label(bias_str))
        rcolor = risk_color(bias_str)
        sid = r["股票代碼"]
        # Goodinfo K線圖連結
        kline_url = "https://goodinfo.tw/tw/StockDetail.asp?STOCK_ID="+sid

        row_data = [sid,names.get(sid,"-"),labels.get(sig,sig),
            r["評分"],r["量比"],bias_str,rlabel,
            r["突破幅度"],r["K棒品質"],r["RS20"],r["高檔反轉"],r["Gate"],r["收盤價"]]
        for i,val in enumerate(row_data,1):
            if i==3:
                make_cell(ws1,row_num,i,val,fill_color=bg,font_color=sig_colors.get(sig,"333333"),bold=True,border=border)
            elif i==7:
                make_cell(ws1,row_num,i,val,fill_color=bg,font_color=rcolor,bold=True,border=border)
            elif i==11 and str(val)=="是":
                make_cell(ws1,row_num,i,val,fill_color=bg,font_color="C00000",bold=True,border=border)
            else:
                make_cell(ws1,row_num,i,val,fill_color=bg,border=border)
        # K線圖超連結欄
        make_cell(ws1,row_num,14,"📊 查看K線",fill_color=bg,border=border,hyperlink=kline_url)
        ws1.row_dimensions[row_num].height = 18; row_num += 1

    ws1.freeze_panes = "A5"
    ws1.auto_filter.ref = "A4:N4"

    # ── Sheet 2：昨日追蹤 ──────────────────────────────
    ws2 = wb.create_sheet("昨日Top5追蹤")
    history = load_top5_history()
    hist_date = None
    for i in range(1, 8):
        d = (datetime.today()-timedelta(days=i)).strftime("%Y-%m-%d")
        if d in history: hist_date = d; break

    ws2.merge_cells("A1:I1")
    make_cell(ws2,1,1,"📊 昨日Top5選股追蹤　（分析日期："+today_str+"）",
              fill_color="1F497D",font_color="FFFFFF",bold=True,size=14)
    ws2.row_dimensions[1].height = 28

    if hist_date and history.get(hist_date):
        prev_top5 = history[hist_date]
        ws2.merge_cells("A2:I2")
        make_cell(ws2,2,1,"推薦日期："+hist_date+"　共 "+str(len(prev_top5))+" 支",
                  fill_color="EBF3FB",bold=True,size=11)
        ws2.row_dimensions[2].height = 22
        h2 = ["代碼","股票名稱","推薦訊號","推薦評分","推薦收盤","今日收盤","漲跌幅","今日診斷","診斷說明"]
        cw2 = [8,12,12,8,8,8,8,12,45]
        for i,(h,w) in enumerate(zip(h2,cw2),1):
            make_cell(ws2,4,i,h,fill_color="2E75B6",font_color="FFFFFF",bold=True,border=border)
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.row_dimensions[4].height = 20
        row_num2 = 5
        for item in prev_top5:
            sid   = item["股票代碼"]
            name  = names.get(sid, item.get("股票名稱","-"))
            status, reason = diagnose_stock(sid, pc, item)
            df_now = pc.get(sid)
            today_close = "-"; chg_str = "-"; chg_color = "333333"
            if df_now is not None and not df_now.empty:
                df_now = df_now.copy()
                df_now["close"] = pd.to_numeric(df_now["close"], errors="coerce")
                tc_val = df_now["close"].iloc[-1]
                today_close = round(tc_val, 2)
                prev_close = item.get("今日收盤", item.get("收盤價", 0))
                if prev_close and prev_close > 0:
                    chg = (tc_val - prev_close) / prev_close * 100
                    chg_str = str(round(chg,1))+"%" if chg < 0 else "+"+str(round(chg,1))+"%"
                    chg_color = "C00000" if chg > 0 else ("217346" if chg < 0 else "333333")
            bg = "FFFFFF" if row_num2%2==0 else "F5F5F5"
            if "✅" in status: sc = "217346"
            elif "❌" in status or "🔴" in status: sc = "C00000"
            elif "⚠️" in status: sc = "E36C09"
            else: sc = "444444"
            row_vals = [sid, name, labels.get(item.get("訊號",""),item.get("訊號","-")),
                item.get("評分","-"), item.get("今日收盤", item.get("收盤價","-")),
                today_close, chg_str, status, reason]
            for i,val in enumerate(row_vals,1):
                if i==7: make_cell(ws2,row_num2,i,val,fill_color=bg,font_color=chg_color,bold=True,border=border)
                elif i==8: make_cell(ws2,row_num2,i,val,fill_color=bg,font_color=sc,bold=True,border=border)
                elif i==9: make_cell(ws2,row_num2,i,val,fill_color=bg,font_color="444444",align="left",border=border)
                else: make_cell(ws2,row_num2,i,val,fill_color=bg,border=border)
            ws2.row_dimensions[row_num2].height = 30; row_num2 += 1
    else:
        ws2.merge_cells("A3:I3")
        make_cell(ws2,3,1,"尚無昨日推薦記錄（首次執行）",fill_color="FFF2CC",font_color="7F6000",size=12)
    ws2.freeze_panes = "A5"

    # 儲存今日Top5
    top5 = []
    for r in results[:5]:
        top5.append({
            "股票代碼": r["股票代碼"], "股票名稱": names.get(r["股票代碼"],"-"),
            "訊號": r["訊號"], "評分": r["評分"],
            "今日收盤": r["收盤價"], "突破點": r.get("突破點", r["收盤價"]),
            "量比": r["量比"],
        })
    history[datetime.today().strftime("%Y-%m-%d")] = top5
    all_dates = sorted(history.keys(), reverse=True)
    for old_d in all_dates[30:]: del history[old_d]
    save_top5_history(history)

    output_path = output_path.replace(".docx",".xlsx")
    wb.save(output_path)
    print("✅ Excel報告已產生："+output_path)
    return output_path

def main():
    print("="*55)
    print("  台股飆股選股程式 v3（證交所+櫃買版）")
    print("  執行時間："+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print("="*55)
    sd = get_stock_list()
    pc, idx, meta = load_cache()
    if not meta["last_update"]: pc, idx, meta = full_init(sd)
    else: pc, idx = incremental_update(pc, idx, meta, sd)
    idf = idx.get("TAIEX")
    print("\n🔍 開始篩選...")
    results = []
    for sid, info in sd.items():
        try:
            r = evaluate_stock(sid, pc.get(sid), idf)
            if r:
                results.append(r)
                print("  "+sid+" "+info["name"]+" | "+r["訊號"]+" | 分="+str(r["評分"])+" | "+r["風險警示"])
        except: pass
    print("\n"+"="*55)
    print("✅ 篩選完成！共 "+str(len(results))+" 支")
    print("="*55)
    ds = datetime.today().strftime("%Y%m%d")
    op = os.path.join(OUTPUT_DIR, "飆股日報_"+ds+".xlsx")
    generate_excel(results, sd, pc, op)
    generate_html(results, sd, pc, op)
    if results:
        df = pd.DataFrame(results)
        # 建立完整 name_map（stock_list 過濾空字串 + industries fallback）
        name_map = {s: i["name"] for s, i in sd.items()
                    if i.get("name") and str(i["name"]).strip()}
        try:
            from sector_analyzer import fetch_all_industries
            ind = fetch_all_industries()
            for sid_k, info in ind.items():
                existing = name_map.get(sid_k, "")
                if (not existing or not str(existing).strip()) and info.get("name"):
                    name_map[sid_k] = info["name"]
        except Exception:
            pass
        df["股票代碼"] = df["股票代碼"].astype(str).str.zfill(4)   # 確保 4 位數字串
        df.insert(1, "股票名稱", df["股票代碼"].map(name_map).fillna("-"))
        df.to_csv(os.path.join(OUTPUT_DIR,"飆股日報_"+ds+".csv"),index=False,encoding="utf-8-sig")
        print(f"💾 CSV已存檔（name_map 含 {len(name_map)} 支）")

if __name__=="__main__": main()
